import hashlib
import io
import os
import re
import secrets
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Cookie, Header, HTTPException, Request, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from passlib.context import CryptContext

from lib.db import db
from models.cms import (
    AdminLogin,
    AdminAccount,
    AdminAccountCreate,
    AdminAccountUpdate,
    AdminUser,
    AuditLog,
    AnalyticsPoint,
    AnalyticsSlice,
    CMSItem,
    CMSItemCreate,
    CMSItemUpdate,
    Lead,
    LeadCreate,
    LeadNote,
    LeadNoteCreate,
    TimelineEvent,
    WhatsAppActionCreate,
    WhatsAppActionResponse,
    WhatsAppTemplate,
    WhatsAppTemplateUpdate,
    LeadUpdate,
    MessageResponse,
    PPDBAnalytics,
    ReportOverview,
    ReportRun,
    ResourceType,
)
from lib.reports import build_weekly_summary, create_report_run, run_scheduled_report_if_due, next_report_time


router = APIRouter()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SESSION_COOKIE = "school_admin_session"
CONTENT_ROLES = {
    "super_admin": {"news", "agenda", "gallery", "major"},
    "content_editor": {"news", "gallery", "major"},
    "agenda_manager": {"agenda"},
    "ppdb_officer": set(),
}


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def normalize_datetime(value: datetime) -> datetime:
    return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value


def serialize_item(document: dict) -> CMSItem:
    document["created_at"] = normalize_datetime(document["created_at"])
    document["updated_at"] = normalize_datetime(document["updated_at"])
    return CMSItem(**document)


def serialize_lead(document: dict) -> Lead:
    document["created_at"] = normalize_datetime(document["created_at"])
    age_hours = max(0, int((utc_now() - document["created_at"]).total_seconds() // 3600))
    document["age_hours"] = age_hours
    document["duplicate_count"] = len(document.get("duplicate_ids", []))
    document["sla_level"] = "critical" if document.get("status") == "new" and age_hours >= 48 else "warning" if document.get("status") == "new" and age_hours >= 24 else "ok"
    if document.get("last_contact_at"):
        document["last_contact_at"] = normalize_datetime(document["last_contact_at"])
    return Lead(**document)


def serialize_admin(document: dict) -> AdminAccount:
    if "created_at" in document:
        document["created_at"] = normalize_datetime(document["created_at"])
    return AdminAccount(**document)


def ensure_content_permission(admin: dict, resource: ResourceType) -> None:
    if resource not in CONTENT_ROLES.get(admin.get("role", ""), set()):
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses untuk konten ini")


def ensure_leads_permission(admin: dict) -> None:
    if admin.get("role") not in {"super_admin", "ppdb_officer"}:
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke data pendaftar")


def ensure_super_admin(admin: dict) -> None:
    if admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Hanya Super Admin yang dapat mengelola akun")


def build_lead_query(kind: str | None, status: str | None, start_date: str | None, end_date: str | None, source: str | None = None) -> dict:
    query: dict = {}
    if kind:
        query["kind"] = kind
    if status:
        query["status"] = status
    if source:
        query["source"] = source
    if start_date or end_date:
        date_query: dict = {}
        try:
            if start_date:
                date_query["$gte"] = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            if end_date:
                date_query["$lt"] = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail="Format tanggal harus YYYY-MM-DD") from exc
        query["created_at"] = date_query
    return query


async def write_audit(admin: dict, action: str, entity_type: str, entity_id: str, summary: str, details: dict | None = None) -> None:
    log = AuditLog(
        actor_id=admin["id"],
        actor_name=admin["name"],
        actor_email=admin["email"],
        actor_role=admin["role"],
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        summary=summary,
        details=details or {},
    )
    await db.audit_logs.insert_one(log.model_dump())


def serialize_audit(document: dict) -> AuditLog:
    document["created_at"] = normalize_datetime(document["created_at"])
    return AuditLog(**document)


async def require_admin(session_token: str | None) -> dict:
    if not session_token:
        raise HTTPException(status_code=401, detail="Sesi admin diperlukan")
    token_hash = hashlib.sha256(session_token.encode()).hexdigest()
    session = await db.admin_sessions.find_one({"token_hash": token_hash})
    if not session or normalize_datetime(session["expires_at"]) <= utc_now():
        raise HTTPException(status_code=401, detail="Sesi admin tidak valid")
    admin = await db.admins.find_one({"id": session["admin_id"]})
    if not admin or not admin.get("is_active", True):
        raise HTTPException(status_code=401, detail="Admin tidak ditemukan")
    return admin


@router.post("/auth/login", response_model=AdminUser)
async def login(payload: AdminLogin, request: Request, response: Response):
    admin = await db.admins.find_one({"email": payload.email.lower().strip()})
    if not admin or not pwd_context.verify(payload.password, admin["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    token = secrets.token_urlsafe(32)
    await db.admin_sessions.insert_one({
        "token_hash": hashlib.sha256(token.encode()).hexdigest(),
        "admin_id": admin["id"],
        "expires_at": utc_now() + timedelta(days=7),
    })
    is_secure = request.headers.get("x-forwarded-proto", request.url.scheme) == "https"
    # "none" is for split-domain deploys (e.g. Vercel frontend calling Railway backend
    # directly instead of through a same-origin proxy rewrite) — browsers reject
    # SameSite=None without Secure, so force secure whenever that's configured.
    samesite = os.environ.get("COOKIE_SAMESITE", "lax").lower()
    if samesite == "none":
        is_secure = True
    response.set_cookie(SESSION_COOKIE, token, httponly=True, secure=is_secure, samesite=samesite, max_age=604800, path="/")
    return AdminUser(id=admin["id"], email=admin["email"], name=admin["name"], role=admin.get("role", "super_admin"))


@router.get("/auth/me", response_model=AdminUser)
async def me(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    return AdminUser(id=admin["id"], email=admin["email"], name=admin["name"], role=admin.get("role", "super_admin"))


@router.post("/auth/logout", response_model=MessageResponse)
async def logout(response: Response, school_admin_session: str | None = Cookie(default=None)):
    if school_admin_session:
        token_hash = hashlib.sha256(school_admin_session.encode()).hexdigest()
        await db.admin_sessions.delete_one({"token_hash": token_hash})
    response.delete_cookie(SESSION_COOKIE, path="/", samesite=os.environ.get("COOKIE_SAMESITE", "lax").lower())
    return MessageResponse(message="Berhasil keluar")


@router.get("/content/{resource}", response_model=list[CMSItem])
async def public_content(resource: ResourceType):
    documents = await db.cms_items.find({"resource": resource, "is_published": True}, {"_id": 0}).sort("date", 1).to_list(200)
    return [serialize_item(document) for document in documents]


@router.get("/admin/content/{resource}", response_model=list[CMSItem])
async def admin_content(resource: ResourceType, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_content_permission(admin, resource)
    documents = await db.cms_items.find({"resource": resource}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [serialize_item(document) for document in documents]


@router.post("/admin/content", response_model=CMSItem, status_code=201)
async def create_content(payload: CMSItemCreate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_content_permission(admin, payload.resource)
    item = CMSItem(**payload.model_dump())
    await db.cms_items.insert_one(item.model_dump())
    await write_audit(admin, "content_created", "content", item.id, f"Membuat {payload.resource}: {payload.title}", {"resource": payload.resource})
    return item


@router.patch("/admin/content/{item_id}", response_model=CMSItem)
async def update_content(item_id: str, payload: CMSItemUpdate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    existing = await db.cms_items.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Konten tidak ditemukan")
    ensure_content_permission(admin, existing["resource"])
    changes = payload.model_dump(exclude_unset=True)
    changes["updated_at"] = utc_now()
    document = await db.cms_items.find_one_and_update({"id": item_id}, {"$set": changes}, return_document=True, projection={"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Konten tidak ditemukan")
    await write_audit(admin, "content_updated", "content", item_id, f"Mengubah {existing['resource']}: {existing['title']}", {"fields": list(changes.keys())})
    return serialize_item(document)


@router.delete("/admin/content/{item_id}", response_model=MessageResponse)
async def delete_content(item_id: str, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    existing = await db.cms_items.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Konten tidak ditemukan")
    ensure_content_permission(admin, existing["resource"])
    result = await db.cms_items.delete_one({"id": item_id})
    if not result.deleted_count:
        raise HTTPException(status_code=404, detail="Konten tidak ditemukan")
    await write_audit(admin, "content_deleted", "content", item_id, f"Menghapus {existing['resource']}: {existing['title']}")
    return MessageResponse(message="Konten dihapus")


@router.post("/leads", response_model=Lead, status_code=201)
async def create_lead(payload: LeadCreate):
    digits = "".join(character for character in payload.phone if character.isdigit())
    normalized_phone = f"62{digits[1:]}" if digits.startswith("0") else digits
    previous = await db.leads.find({"normalized_phone": normalized_phone}, {"id": 1, "_id": 0}).to_list(100)
    previous_ids = [item["id"] for item in previous]
    lead = Lead(**payload.model_dump(), normalized_phone=normalized_phone, duplicate_ids=previous_ids)
    lead.last_contact_at = lead.created_at
    await db.leads.insert_one(lead.model_dump())
    if previous_ids:
        await db.leads.update_many({"id": {"$in": previous_ids}}, {"$addToSet": {"duplicate_ids": lead.id}})
    return lead


@router.get("/admin/leads", response_model=list[Lead])
async def list_leads(kind: str | None = None, status: str | None = None, start_date: str | None = None, end_date: str | None = None, source: str | None = None, scope: str = "all", school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    query = build_lead_query(kind, status, start_date, end_date, source)
    if admin["role"] == "ppdb_officer":
        if scope == "mine":
            query["assigned_to_id"] = admin["id"]
        elif scope == "unassigned":
            query["assigned_to_id"] = None
        else:
            query["$or"] = [{"assigned_to_id": admin["id"]}, {"assigned_to_id": None}]
    elif scope == "mine":
        query["assigned_to_id"] = admin["id"]
    elif scope == "unassigned":
        query["assigned_to_id"] = None
    documents = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_lead(document) for document in documents]


@router.get("/admin/leads/export.xlsx")
async def export_leads(kind: str | None = None, status: str | None = None, start_date: str | None = None, end_date: str | None = None, source: str | None = None, scope: str = "all", template: str = "full", school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    if template not in {"full", "compact", "contacts"}:
        raise HTTPException(status_code=422, detail="Template ekspor tidak valid")
    query = build_lead_query(kind, status, start_date, end_date, source)
    if admin["role"] == "ppdb_officer":
        if scope == "mine":
            query["assigned_to_id"] = admin["id"]
        elif scope == "unassigned":
            query["assigned_to_id"] = None
        else:
            query["$or"] = [{"assigned_to_id": admin["id"]}, {"assigned_to_id": None}]
    elif scope == "mine":
        query["assigned_to_id"] = admin["id"]
    elif scope == "unassigned":
        query["assigned_to_id"] = None
    documents = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    leads = [serialize_lead(document) for document in documents]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data Leads"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "SMK TERATAI PUTIH GLOBAL 2 BEKASI"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0A3358")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = "Jl. Rajawali V Perumnas 1, Kayuringin Jaya, Bekasi Selatan · reportterput2@gmail.com"
    sheet.merge_cells("A3:J3")
    sheet["A3"] = f"Template: {template.title()} · Dicetak {utc_now().strftime('%d-%m-%Y %H:%M UTC')}"
    column_sets = {
        "full": ["ID", "Jenis", "Nama", "WhatsApp", "Jurusan", "Pertanyaan", "Sumber", "Petugas", "Status", "Waktu Masuk"],
        "compact": ["Nama", "WhatsApp", "Jurusan", "Sumber", "Status", "Petugas"],
        "contacts": ["Nama", "WhatsApp", "Jenis", "Jurusan"],
    }
    headers = column_sets[template]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0F4C81")
    header_row = 4
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for lead in leads:
        full_row = {"ID": lead.id, "Jenis": lead.kind, "Nama": lead.name, "WhatsApp": lead.phone, "Jurusan": lead.major or "", "Pertanyaan": lead.question or "", "Sumber": lead.source, "Petugas": lead.assigned_to_name or "Belum ditugaskan", "Status": lead.status, "Waktu Masuk": lead.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}
        sheet.append([full_row[header] for header in headers])
    widths = [38, 14, 25, 18, 20, 45, 16, 24, 16, 22][:len(headers)]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    summary = workbook.create_sheet("Ringkasan")
    summary.append(["Ringkasan Ekspor Leads", "Jumlah"])
    summary.append(["Total data", len(leads)])
    summary.append(["Pendaftar SPMB", sum(lead.kind == "ppdb" for lead in leads)])
    summary.append(["Pertanyaan kontak", sum(lead.kind == "contact" for lead in leads)])
    summary.append(["Status baru", sum(lead.status == "new" for lead in leads)])
    summary.append(["Tindak lanjut", sum(lead.status == "follow_up" for lead in leads)])
    summary.append(["Selesai", sum(lead.status == "done" for lead in leads)])
    summary.append(["Filter jenis", kind or "Semua"])
    summary.append(["Filter status", status or "Semua"])
    summary.append(["Filter sumber", source or "Semua"])
    summary.append(["Template", template.title()])
    summary.append(["Rentang tanggal", f"{start_date or '-'} s/d {end_date or '-'}"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 28
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"leads-teratai-{utc_now().date().isoformat()}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.patch("/admin/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, payload: LeadUpdate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    existing = await db.leads.find_one({"id": lead_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    changes = payload.model_dump(exclude_unset=True)
    if "assigned_to_id" in changes:
        ensure_super_admin(admin)
        assignee_id = changes["assigned_to_id"]
        if assignee_id:
            assignee = await db.admins.find_one({"id": assignee_id, "role": "ppdb_officer", "is_active": True})
            if not assignee:
                raise HTTPException(status_code=422, detail="Petugas SPMB aktif tidak ditemukan")
            changes["assigned_to_name"] = assignee["name"]
        else:
            changes["assigned_to_name"] = None
    if admin["role"] == "ppdb_officer" and existing.get("assigned_to_id") not in {None, admin["id"]}:
        raise HTTPException(status_code=403, detail="Lead ini ditugaskan kepada petugas lain")
    document = await db.leads.find_one_and_update({"id": lead_id}, {"$set": changes}, return_document=True, projection={"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    if "assigned_to_id" in changes:
        await db.leads.update_one({"id": lead_id}, {"$set": {"last_contact_type": "Penugasan petugas", "last_contact_at": utc_now(), "last_contact_by": admin["name"]}})
        await write_audit(admin, "lead_assigned", "lead", lead_id, f"Menugaskan lead {existing['name']} kepada {changes.get('assigned_to_name') or 'belum ditugaskan'}", {"assigned_to_id": changes.get("assigned_to_id")})
    if "status" in changes:
        await db.leads.update_one({"id": lead_id}, {"$set": {"last_contact_type": "Perubahan status", "last_contact_at": utc_now(), "last_contact_by": admin["name"]}})
        await write_audit(admin, "lead_status_updated", "lead", lead_id, f"Mengubah status lead {existing['name']} menjadi {changes['status']}", {"before": existing.get("status"), "after": changes["status"]})
    return serialize_lead(document)


@router.get("/admin/leads/{lead_id}/duplicates", response_model=list[Lead])
async def lead_duplicates(lead_id: str, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    lead = await db.leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    documents = await db.leads.find({"normalized_phone": lead.get("normalized_phone")}, {"_id": 0}).sort("created_at", 1).to_list(100)
    return [serialize_lead(document) for document in documents]


@router.get("/admin/leads/{lead_id}/notes", response_model=list[LeadNote])
async def list_lead_notes(lead_id: str, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    documents = await db.lead_notes.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for document in documents:
        document["created_at"] = normalize_datetime(document["created_at"])
    return [LeadNote(**document) for document in documents]


@router.post("/admin/leads/{lead_id}/notes", response_model=LeadNote, status_code=201)
async def add_lead_note(lead_id: str, payload: LeadNoteCreate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    lead = await db.leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    if admin["role"] == "ppdb_officer" and lead.get("assigned_to_id") not in {None, admin["id"]}:
        raise HTTPException(status_code=403, detail="Lead ini ditugaskan kepada petugas lain")
    note = LeadNote(lead_id=lead_id, author_id=admin["id"], author_name=admin["name"], **payload.model_dump())
    await db.lead_notes.insert_one(note.model_dump())
    await db.leads.update_one({"id": lead_id}, {"$set": {"last_contact_type": "Catatan petugas", "last_contact_at": utc_now(), "last_contact_by": admin["name"], "next_action_date": payload.next_action_date}})
    await write_audit(admin, "lead_note_added", "lead", lead_id, f"Menambahkan catatan untuk {lead['name']}")
    return note


@router.get("/admin/leads/{lead_id}/timeline", response_model=list[TimelineEvent])
async def lead_timeline(lead_id: str, types: str | None = None, days: int | None = None, start_date: str | None = None, end_date: str | None = None, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    lead = await db.leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    if admin["role"] == "ppdb_officer" and lead.get("assigned_to_id") not in {None, admin["id"]}:
        raise HTTPException(status_code=403, detail="Lead ini ditugaskan kepada petugas lain")
    events = [TimelineEvent(id=f"created-{lead_id}", event_type="created", title="Lead dibuat", description=f"Pendaftar masuk dari {lead.get('source', 'website')}", actor_name="Sistem", created_at=normalize_datetime(lead["created_at"]), metadata={"source": lead.get("source", "website")})]
    audits = await db.audit_logs.find({"entity_id": lead_id, "action": {"$in": ["lead_assigned", "lead_status_updated"]}}, {"_id": 0}).to_list(500)
    for item in audits:
        events.append(TimelineEvent(id=item["id"], event_type="assignment" if item["action"] == "lead_assigned" else "status", title="Penugasan petugas" if item["action"] == "lead_assigned" else "Status diperbarui", description=item["summary"], actor_name=item["actor_name"], created_at=normalize_datetime(item["created_at"]), metadata=item.get("details", {})))
    notes = await db.lead_notes.find({"lead_id": lead_id}, {"_id": 0}).to_list(500)
    for item in notes:
        events.append(TimelineEvent(id=item["id"], event_type="note", title="Catatan petugas", description=item["text"], actor_name=item["author_name"], created_at=normalize_datetime(item["created_at"]), metadata={"next_action_date": item.get("next_action_date")}))
    communications = await db.communication_events.find({"lead_id": lead_id}, {"_id": 0}).to_list(500)
    for item in communications:
        events.append(TimelineEvent(id=item["id"], event_type="whatsapp", title="Pesan WhatsApp dibuka", description=item["message"], actor_name=item["actor_name"], created_at=normalize_datetime(item["created_at"]), metadata={"template": item["template"]}))
    if types:
        allowed_types = set(types.split(","))
        events = [event for event in events if event.event_type in allowed_types]
    if days is not None:
        if days not in {7, 30, 90}:
            raise HTTPException(status_code=422, detail="Periode harus 7, 30, atau 90 hari")
        threshold = utc_now() - timedelta(days=days)
        events = [event for event in events if event.created_at >= threshold]
    if start_date or end_date:
        date_filter = build_lead_query(None, None, start_date, end_date).get("created_at", {})
        if "$gte" in date_filter:
            events = [event for event in events if event.created_at >= date_filter["$gte"]]
        if "$lt" in date_filter:
            events = [event for event in events if event.created_at < date_filter["$lt"]]
    return sorted(events, key=lambda item: item.created_at, reverse=True)


@router.post("/admin/leads/{lead_id}/whatsapp", response_model=WhatsAppActionResponse, status_code=201)
async def open_whatsapp_action(lead_id: str, payload: WhatsAppActionCreate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    lead = await db.leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    if admin["role"] == "ppdb_officer" and lead.get("assigned_to_id") not in {None, admin["id"]}:
        raise HTTPException(status_code=403, detail="Lead ini ditugaskan kepada petugas lain")
    template = await db.whatsapp_templates.find_one({"key": payload.template, "is_active": True}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=422, detail="Template WhatsApp tidak aktif")
    variables = {"{nama}": lead["name"], "{jurusan}": lead.get("major") or "program pilihan", "{petugas}": admin["name"], "{sekolah}": "SMK Teratai Putih Global 2 Bekasi"}
    message = template["content"]
    for variable, value in variables.items():
        message = message.replace(variable, value)
    phone = lead.get("normalized_phone") or "".join(character for character in lead["phone"] if character.isdigit())
    event = {"id": str(uuid4()), "lead_id": lead_id, "template": payload.template, "message": message, "url": f"https://wa.me/{phone}?text={quote(message)}", "actor_id": admin["id"], "actor_name": admin["name"], "created_at": utc_now()}
    await db.communication_events.insert_one(event)
    await db.leads.update_one({"id": lead_id}, {"$set": {"last_contact_type": f"WhatsApp · {template['label']}", "last_contact_at": utc_now(), "last_contact_by": admin["name"]}})
    await write_audit(admin, "whatsapp_opened", "lead", lead_id, f"Membuka pesan WhatsApp {payload.template} untuk {lead['name']}", {"template": payload.template})
    return WhatsAppActionResponse(**event)


@router.get("/admin/whatsapp-templates", response_model=list[WhatsAppTemplate])
async def list_whatsapp_templates(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    documents = await db.whatsapp_templates.find({}, {"_id": 0}).sort("order", 1).to_list(20)
    for document in documents:
        document["updated_at"] = normalize_datetime(document["updated_at"])
    return [WhatsAppTemplate(**document) for document in documents]


@router.patch("/admin/whatsapp-templates/{template_key}", response_model=WhatsAppTemplate)
async def update_whatsapp_template(template_key: str, payload: WhatsAppTemplateUpdate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    changes = payload.model_dump(exclude_unset=True)
    if payload.content is not None:
        variables = set(re.findall(r"\{[^{}]+\}", payload.content))
        unsupported = variables - {"{nama}", "{jurusan}", "{petugas}", "{sekolah}"}
        if unsupported:
            raise HTTPException(status_code=422, detail=f"Variabel tidak didukung: {', '.join(sorted(unsupported))}")
    changes.update({"updated_by": admin["name"], "updated_at": utc_now()})
    document = await db.whatsapp_templates.find_one_and_update({"key": template_key}, {"$set": changes}, return_document=True, projection={"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    await write_audit(admin, "whatsapp_template_updated", "admin", template_key, f"Mengubah template WhatsApp {document['label']}")
    return WhatsAppTemplate(**document)


@router.post("/admin/whatsapp-templates/{template_key}/reset", response_model=WhatsAppTemplate)
async def reset_whatsapp_template(template_key: str, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    defaults = {
        "greeting": "Halo {nama}, kami dari {sekolah}. Terima kasih sudah mendaftar pada jurusan {jurusan}. Saya {petugas}, apakah ada informasi yang dapat kami bantu?",
        "documents": "Halo {nama}, kami mengingatkan kelengkapan berkas SPMB {sekolah} untuk jurusan {jurusan}. Mohon konfirmasi jika berkas sudah siap.",
        "visit": "Halo {nama}, kami mengundang Anda untuk mengatur jadwal kunjungan ke {sekolah}. Silakan balas dengan waktu yang paling sesuai.",
        "final_follow_up": "Halo {nama}, kami menindaklanjuti kembali minat pendaftaran jurusan {jurusan} di {sekolah}. Apakah proses pendaftaran ingin dilanjutkan?",
    }
    if template_key not in defaults:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    document = await db.whatsapp_templates.find_one_and_update({"key": template_key}, {"$set": {"content": defaults[template_key], "is_active": True, "updated_by": admin["name"], "updated_at": utc_now()}}, return_document=True, projection={"_id": 0})
    await write_audit(admin, "whatsapp_template_reset", "admin", template_key, f"Memulihkan template WhatsApp {template_key}")
    return WhatsAppTemplate(**document)


@router.get("/admin/reports", response_model=ReportOverview)
async def report_overview(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    summary = await build_weekly_summary()
    documents = await db.report_runs.find({}, {"_id": 0}).sort("created_at", -1).to_list(20)
    for document in documents:
        document["created_at"] = normalize_datetime(document["created_at"])
    return ReportOverview(
        recipient=os.environ.get("REPORT_RECIPIENT", "pimpinan@example.com"),
        sender=os.environ.get("REPORT_SENDER", "onboarding@resend.dev"),
        delivery_mode="simulated",
        schedule="Setiap Senin, 07.00 WIB",
        next_run=next_report_time().astimezone(timezone.utc),
        preview=summary,
        runs=[ReportRun(**document) for document in documents],
    )


@router.post("/admin/reports/run", response_model=ReportRun, status_code=201)
async def run_report_simulation(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    document = await create_report_run("manual")
    await write_audit(admin, "report_simulated", "admin", admin["id"], "Menjalankan simulasi laporan mingguan SPMB")
    return ReportRun(**document)


@router.post("/cron/weekly-report")
@router.get("/cron/weekly-report")
async def trigger_weekly_report(authorization: str | None = Header(default=None)):
    """Entry point for an external scheduler (e.g. Vercel Cron) since serverless deployments
    can't run the in-process report_scheduler background loop. Idempotent per ISO week.
    Registered as both GET and POST: Vercel Cron always invokes with GET, but the endpoint
    also accepts POST for manual/other-scheduler triggers."""
    cron_secret = os.environ.get("CRON_SECRET")
    if not cron_secret or authorization != f"Bearer {cron_secret}":
        raise HTTPException(status_code=401, detail="Unauthorized")
    document = await run_scheduled_report_if_due()
    if document is None:
        return {"triggered": False}
    return {"triggered": True, "run_id": document["id"]}


@router.get("/admin/assignees", response_model=list[AdminUser])
async def list_assignees(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    documents = await db.admins.find({"role": "ppdb_officer", "is_active": True}, {"_id": 0}).sort("name", 1).to_list(100)
    return [AdminUser(id=item["id"], email=item["email"], name=item["name"], role=item["role"]) for item in documents]


@router.get("/admin/audit", response_model=list[AuditLog])
async def list_audit(actor_id: str | None = None, action: str | None = None, start_date: str | None = None, end_date: str | None = None, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    query: dict = {}
    if actor_id:
        query["actor_id"] = actor_id
    if action:
        query["action"] = action
    if start_date or end_date:
        query["created_at"] = build_lead_query(None, None, start_date, end_date).get("created_at", {})
    documents = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [serialize_audit(document) for document in documents]


@router.get("/admin/analytics", response_model=PPDBAnalytics)
async def ppdb_analytics(days: int = 30, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_leads_permission(admin)
    if days not in {30, 90, 365}:
        raise HTTPException(status_code=422, detail="Periode harus 30, 90, atau 365 hari")
    start = utc_now() - timedelta(days=days)
    documents = await db.leads.find({"kind": "ppdb", "created_at": {"$gte": start}}, {"_id": 0}).to_list(10000)
    leads = [serialize_lead(document) for document in documents]
    major_counts: dict[str, int] = {}
    source_counts = {key: 0 for key in ["website", "whatsapp", "instagram", "walk_in", "referral"]}
    week_counts: dict[str, int] = {}
    for lead in leads:
        major_counts[lead.major or "Belum memilih"] = major_counts.get(lead.major or "Belum memilih", 0) + 1
        source_counts[lead.source] = source_counts.get(lead.source, 0) + 1
        week_start = (lead.created_at - timedelta(days=lead.created_at.weekday())).date().isoformat()
        week_counts[week_start] = week_counts.get(week_start, 0) + 1
    return PPDBAnalytics(
        period_days=days,
        total=len(leads),
        new_count=sum(lead.status == "new" for lead in leads),
        follow_up_count=sum(lead.status == "follow_up" for lead in leads),
        done_count=sum(lead.status == "done" for lead in leads),
        by_major=[AnalyticsSlice(label=label, value=value) for label, value in sorted(major_counts.items(), key=lambda item: item[1], reverse=True)],
        by_source=[AnalyticsSlice(label=label, value=value) for label, value in source_counts.items()],
        weekly=[AnalyticsPoint(label=label, count=value) for label, value in sorted(week_counts.items())],
    )


@router.get("/admin/users", response_model=list[AdminAccount])
async def list_admins(school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    documents = await db.admins.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(100)
    return [serialize_admin(document) for document in documents]


@router.post("/admin/users", response_model=AdminAccount, status_code=201)
async def create_admin(payload: AdminAccountCreate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    email = payload.email.lower().strip()
    if await db.admins.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email admin sudah digunakan")
    account = AdminAccount(email=email, name=payload.name, role=payload.role)
    document = account.model_dump()
    document["password_hash"] = pwd_context.hash(payload.password)
    await db.admins.insert_one(document)
    await write_audit(admin, "admin_created", "admin", account.id, f"Membuat akun {account.name} sebagai {account.role}", {"email": account.email, "role": account.role})
    return account


@router.patch("/admin/users/{admin_id}", response_model=AdminAccount)
async def update_admin(admin_id: str, payload: AdminAccountUpdate, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    changes = payload.model_dump(exclude_unset=True, exclude={"password"})
    if payload.password:
        changes["password_hash"] = pwd_context.hash(payload.password)
    document = await db.admins.find_one_and_update({"id": admin_id}, {"$set": changes}, return_document=True, projection={"_id": 0, "password_hash": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Akun staf tidak ditemukan")
    await write_audit(admin, "admin_updated", "admin", admin_id, f"Mengubah akun {document['name']}", {"fields": list(changes.keys())})
    return serialize_admin(document)


@router.delete("/admin/users/{admin_id}", response_model=MessageResponse)
async def delete_admin(admin_id: str, school_admin_session: str | None = Cookie(default=None)):
    admin = await require_admin(school_admin_session)
    ensure_super_admin(admin)
    if admin["id"] == admin_id:
        raise HTTPException(status_code=400, detail="Super Admin tidak dapat menghapus akun sendiri")
    existing = await db.admins.find_one({"id": admin_id})
    result = await db.admins.delete_one({"id": admin_id})
    if not result.deleted_count:
        raise HTTPException(status_code=404, detail="Akun staf tidak ditemukan")
    await db.admin_sessions.delete_many({"admin_id": admin_id})
    await write_audit(admin, "admin_deleted", "admin", admin_id, f"Menghapus akun {existing['name'] if existing else admin_id}")
    return MessageResponse(message="Akun staf dihapus")