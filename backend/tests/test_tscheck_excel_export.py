"""Excel export follows filters, has two sheets, and is limited to super_admin/ppdb_officer roles."""

import io
import uuid

from openpyxl import load_workbook

SUPER_ADMIN = {"email": "admin@terataiputih2.sch.id", "password": "TerataiAdmin2026!"}
EDITOR = {"email": "editor@terataiputih2.sch.id", "password": "EditorTeratai2026!"}
AGENDA_MANAGER = {"email": "agenda@terataiputih2.sch.id", "password": "AgendaTeratai2026!"}


def login(client, creds):
    resp = client.post("/auth/login", json=creds)
    assert resp.status_code == 200, resp.text
    return resp.cookies


def test_export_has_two_sheets_and_respects_filter(client):
    cookies = login(client, SUPER_ADMIN)
    suffix = uuid.uuid4().hex[:8]
    unique_name = f"tscheck-export-{suffix}"

    create_resp = client.post(
        "/leads",
        json={"kind": "ppdb", "name": unique_name, "phone": "081200000000", "major": "RPL"},
    )
    assert create_resp.status_code == 201, create_resp.text
    lead_id = create_resp.json()["id"]

    export_resp = client.get("/admin/leads/export.xlsx", params={"kind": "ppdb"}, cookies=cookies)
    assert export_resp.status_code == 200, export_resp.text
    assert export_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(io.BytesIO(export_resp.content))
    assert workbook.sheetnames == ["Data Leads", "Ringkasan"]

    data_sheet = workbook["Data Leads"]
    ids_in_sheet = [row[0] for row in data_sheet.iter_rows(min_row=2, values_only=True)]
    assert lead_id in ids_in_sheet, "created ppdb lead must appear in filtered export"
    kinds_in_sheet = {row[1] for row in data_sheet.iter_rows(min_row=2, values_only=True)}
    assert kinds_in_sheet <= {"ppdb"}, "kind=ppdb filter must exclude contact leads"

    summary_sheet = workbook["Ringkasan"]
    summary_rows = {row[0]: row[1] for row in summary_sheet.iter_rows(min_row=2, values_only=True)}
    assert "Total data" in summary_rows
    assert summary_rows["Total data"] == len(ids_in_sheet)


def test_export_forbidden_for_content_editor_and_agenda_manager(client):
    for creds in (EDITOR, AGENDA_MANAGER):
        cookies = login(client, creds)
        resp = client.get("/admin/leads/export.xlsx", cookies=cookies)
        assert resp.status_code == 403, resp.text
